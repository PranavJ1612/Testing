import openpyxl


class HomePageData:
    test_HomePage_Data = [
        {"emailId": "pranav@gmail.com", "password": "12345", "fullName": "Pranav Jagdale", "gender": "Male"},
        {"emailId": "maitra@gmail.com", "password": "54321", "fullName": "Maitra J", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Pranav\\OneDrive\\Desktop\\excelDemo.xlsx")
        sheet = book.active
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == test_case_name:
                for column in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
        return [Dict]

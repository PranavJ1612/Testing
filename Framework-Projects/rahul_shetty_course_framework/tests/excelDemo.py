import openpyxl

Dict = {}
book = openpyxl.load_workbook("C:\\Users\\Pranav\\OneDrive\\Desktop\\excelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

# Getting & setting values from excel
sheet.cell(row=2, column=2).value = "Pranav"
print(sheet.cell(row=2, column=2).value)
print(sheet['A2'].value)

print(sheet.max_row)
print(sheet.max_column)


for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=row,column=1).value == "Test2":
        for column in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=column).value] = sheet.cell(row=row, column=column).value
            # print(sheet.cell(row=row, column=column).value)


print("Dictionary:",Dict)